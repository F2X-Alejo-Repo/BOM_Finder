"""Unit tests for the XLSX export adapter."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from bom_workbench.domain.entities import BomRow
from bom_workbench.domain.ports import ExportOptions
from bom_workbench.infrastructure.exporters import XlsxExporter


def _make_rows() -> list[BomRow]:
    return [
        BomRow(
            project_id=1,
            source_file="sample.csv",
            designator="R1, R2",
            comment="=1+1",
            footprint="0402",
            lcsc_link="https://lcsc.com/product-detail/example",
            lcsc_part_number="-C12345",
        )
    ]


@pytest.mark.anyio
async def test_procurement_export_writes_exact_contract(tmp_path: Path) -> None:
    output_path = tmp_path / "procurement.xlsx"
    exporter = XlsxExporter()

    result = await exporter.export_procurement_bom(
        _make_rows(),
        output_path,
        ExportOptions(include_metadata_sheet=True),
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["BOM"]

    assert result.output_path == str(output_path)
    assert result.rows_exported == 1
    assert result.sheets_created == ["BOM", "Metadata"]
    assert output_path.exists()
    assert workbook.sheetnames == ["BOM", "Metadata"]
    assert [sheet.cell(row=1, column=index).value for index in range(1, 6)] == [
        "Designator",
        "Comment",
        "Footprint",
        "LCSC LINK",
        "LCSC PART #",
    ]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:E2"
    assert sheet["B2"].value == "'=1+1"
    assert sheet["E2"].value == "'-C12345"
    assert sheet["D2"].hyperlink is not None
    assert sheet["D2"].hyperlink.target == "https://lcsc.com/product-detail/example"
    assert sheet["D2"].font.underline == "single"
    assert any("formula-like text sanitized" in warning for warning in result.warnings)
    assert workbook["Metadata"]["A1"].value == "Export Metadata"


@pytest.mark.anyio
async def test_procurement_export_can_skip_metadata_sheet(tmp_path: Path) -> None:
    output_path = tmp_path / "procurement_no_metadata.xlsx"
    exporter = XlsxExporter()

    result = await exporter.export_procurement_bom(
        _make_rows(),
        output_path,
        ExportOptions(include_metadata_sheet=False),
    )

    workbook = load_workbook(output_path, data_only=False)

    assert workbook.sheetnames == ["BOM"]
    assert result.sheets_created == ["BOM"]


@pytest.mark.anyio
async def test_full_table_export_strips_illegal_worksheet_characters(tmp_path: Path) -> None:
    output_path = tmp_path / "full_table.xlsx"
    exporter = XlsxExporter()
    rows = [
        BomRow(
            project_id=1,
            source_file="sample\x00.csv",
            designator="F1",
            comment="16V 100A 40m\x00 1812 PTC Resettable Fuses RoHS",
            category="Protection\x0bDevices",
            lcsc_part_number="C99999",
        )
    ]

    result = await exporter.export_full_table(
        rows,
        output_path,
        ExportOptions(include_metadata_sheet=True),
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Full Table"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
    comment_column = headers.index("comment") + 1
    category_column = headers.index("category") + 1

    assert result.sheets_created == ["Full Table", "Metadata"]
    assert sheet.cell(row=2, column=comment_column).value == "16V 100A 40m 1812 PTC Resettable Fuses RoHS"
    assert sheet.cell(row=2, column=category_column).value == "ProtectionDevices"
    assert workbook["Metadata"]["B5"].value == "sample.csv"
    assert any("illegal worksheet characters removed" in warning for warning in result.warnings)
