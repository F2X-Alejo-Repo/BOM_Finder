"""Integration tests for the export workflow."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from bom_workbench.application import ExportBomUseCase
from bom_workbench.domain.entities import BomRow
from bom_workbench.domain.ports import ExportOptions
from bom_workbench.infrastructure.exporters import XlsxExporter


@pytest.mark.anyio
async def test_export_use_case_writes_procurement_workbook(tmp_path: Path) -> None:
    rows = [
        BomRow(
            project_id=1,
            source_file="sample.csv",
            designator="C1",
            comment="10uF",
            footprint="0603",
            lcsc_link="https://lcsc.com/product-detail/example",
            lcsc_part_number="C56789",
        )
    ]
    output_path = tmp_path / "export.xlsx"

    use_case = ExportBomUseCase(XlsxExporter())
    result = await use_case.export(
        rows,
        output_path,
        ExportOptions(),
        target="procurement_bom",
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["BOM"]

    assert result.rows_exported == 1
    assert result.output_path == str(output_path)
    assert output_path.exists()
    assert workbook.sheetnames == ["BOM", "Metadata"]
    assert [sheet.cell(row=1, column=index).value for index in range(1, 6)] == [
        "Designator",
        "Comment",
        "Footprint",
        "LCSC LINK",
        "LCSC PART #",
    ]
    assert sheet["A2"].value == "C1"
    assert sheet["D2"].hyperlink is not None
