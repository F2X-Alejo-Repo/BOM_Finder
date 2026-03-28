"""XLSX export adapter."""

from __future__ import annotations

import asyncio
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bom_workbench.domain.entities import BomRow
from bom_workbench.domain.ports import ExportOptions, ExportResult, IExporter

__all__ = ["XlsxExporter"]


@dataclass(slots=True)
class _ExportPayload:
    rows: list[dict[str, Any]]
    headers: list[str]
    sheet_title: str
    metadata_title: str | None
    warnings: list[str]


class XlsxExporter(IExporter):
    """Generate Excel exports with deterministic formatting."""

    _procurement_columns: list[tuple[str, str]] = [
        ("Designator", "designator"),
        ("Comment", "comment"),
        ("Footprint", "footprint"),
        ("LCSC LINK", "lcsc_link"),
        ("LCSC PART #", "lcsc_part_number"),
    ]

    async def export_procurement_bom(
        self,
        rows: Sequence[BomRow],
        output_path: Path,
        options: ExportOptions,
    ) -> ExportResult:
        row_list = list(rows)
        return await asyncio.to_thread(
            self._write_workbook,
            row_list,
            output_path,
            options,
            self._build_procurement_payload(row_list),
        )

    async def export_full_table(
        self,
        rows: Sequence[BomRow],
        output_path: Path,
        options: ExportOptions,
    ) -> ExportResult:
        row_list = list(rows)
        return await asyncio.to_thread(
            self._write_workbook,
            row_list,
            output_path,
            options,
            self._build_full_table_payload(row_list),
        )

    async def export_filtered_view(
        self,
        rows: Sequence[BomRow],
        columns: Sequence[str],
        output_path: Path,
        options: ExportOptions,
    ) -> ExportResult:
        row_list = list(rows)
        return await asyncio.to_thread(
            self._write_workbook,
            row_list,
            output_path,
            options,
            self._build_filtered_payload(row_list, list(columns)),
        )

    def _build_procurement_payload(self, rows: list[BomRow]) -> _ExportPayload:
        payload_rows: list[dict[str, Any]] = []
        warnings: list[str] = []
        for row in rows:
            row_data = self._row_mapping(row)
            payload_rows.append(
                {
                    header: row_data.get(field, "")
                    for header, field in self._procurement_columns
                }
            )
            warnings.extend(self._row_warnings(row_data, self._procurement_columns))
        return _ExportPayload(
            rows=payload_rows,
            headers=[header for header, _field in self._procurement_columns],
            sheet_title="BOM",
            metadata_title="Metadata",
            warnings=warnings,
        )

    def _build_full_table_payload(self, rows: list[BomRow]) -> _ExportPayload:
        field_names = self._exportable_fields(rows)
        payload_rows = []
        warnings: list[str] = []
        for row in rows:
            row_data = self._row_mapping(row)
            payload_rows.append({field: row_data.get(field, "") for field in field_names})
            warnings.extend(self._row_warnings(row_data, [(field, field) for field in field_names]))
        return _ExportPayload(
            rows=payload_rows,
            headers=field_names,
            sheet_title="Full Table",
            metadata_title="Metadata",
            warnings=warnings,
        )

    def _build_filtered_payload(self, rows: list[BomRow], columns: list[str]) -> _ExportPayload:
        warnings: list[str] = []
        payload_rows: list[dict[str, Any]] = []
        for row in rows:
            row_data = self._row_mapping(row)
            payload_rows.append({column: row_data.get(column, "") for column in columns})
            warnings.extend(self._row_warnings(row_data, [(column, column) for column in columns]))
        return _ExportPayload(
            rows=payload_rows,
            headers=columns,
            sheet_title="Filtered View",
            metadata_title="Metadata",
            warnings=warnings,
        )

    def _write_workbook(
        self,
        rows: list[BomRow],
        output_path: Path,
        options: ExportOptions,
        payload: _ExportPayload,
    ) -> ExportResult:
        started = perf_counter()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        sheet = workbook.create_sheet(title=payload.sheet_title)
        self._write_sheet(sheet, payload.rows, payload.headers, options, payload.sheet_title)

        sheets_created = [sheet.title]
        if options.include_metadata_sheet:
            metadata = workbook.create_sheet(title=payload.metadata_title or "Metadata")
            self._write_metadata_sheet(metadata, rows, payload.sheet_title, payload.warnings)
            sheets_created.append(metadata.title)

        workbook.save(output_path)

        duration = perf_counter() - started
        file_size = output_path.stat().st_size if output_path.exists() else 0
        return ExportResult(
            output_path=str(output_path),
            rows_exported=len(rows),
            sheets_created=sheets_created,
            warnings=payload.warnings,
            duration_seconds=duration,
            file_size_bytes=file_size,
        )

    def _write_sheet(
        self,
        sheet,
        rows: list[dict[str, Any]],
        headers: list[str],
        options: ExportOptions,
        sheet_title: str,
    ) -> None:
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10, color="333333")
        hyperlink_font = Font(name="Segoe UI", size=10, color="4A9EFF", underline="single")
        left_alignment = Alignment(horizontal="left", vertical="center")

        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(
                row=1,
                column=column_index,
                value=self._worksheet_safe_value(header),
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_alignment

        for row_index, row in enumerate(rows, start=2):
            for column_index, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_index, column=column_index)
                value = row.get(header, "")
                value = self._sanitize_value(value, options)
                cell.value = value
                cell.alignment = left_alignment
                cell.font = data_font
                if (
                    options.preserve_hyperlinks
                    and isinstance(value, str)
                    and header == "LCSC LINK"
                    and value
                    and not value.startswith("'")
                ):
                    cell.hyperlink = value
                    cell.font = hyperlink_font

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 1)}"
        self._set_column_widths(sheet, rows, headers, sheet_title)

    def _write_metadata_sheet(
        self,
        sheet,
        rows: list[BomRow],
        export_target: str,
        warnings: list[str],
    ) -> None:
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        section_font = Font(name="Segoe UI", size=10, bold=True, color="1A1A2E")
        label_font = Font(name="Segoe UI", size=10, bold=True, color="333333")
        value_font = Font(name="Segoe UI", size=10, color="333333")
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        left_alignment = Alignment(horizontal="left", vertical="center")

        entries = [
            ("Export Metadata", None, header_font, header_fill),
            ("Export timestamp (UTC)", datetime.now(UTC).isoformat(), label_font, None),
            ("Export target", export_target, label_font, None),
            ("Total rows", len(rows), label_font, None),
            ("Source files", self._source_files(rows), label_font, None),
            ("Warnings summary", None, section_font, None),
        ]

        current_row = 1
        for label, value, font, fill in entries:
            safe_label = self._worksheet_safe_value(label)
            sheet.cell(row=current_row, column=1, value=safe_label).font = font
            sheet.cell(row=current_row, column=1).alignment = left_alignment
            if fill is not None:
                sheet.cell(row=current_row, column=1).fill = fill
            if value is not None:
                safe_value = self._worksheet_safe_value(value)
                sheet.cell(row=current_row, column=2, value=safe_value).font = value_font
                sheet.cell(row=current_row, column=2).alignment = left_alignment
            current_row += 1

        if warnings:
            for warning in warnings:
                safe_warning = self._worksheet_safe_value(warning)
                sheet.cell(row=current_row, column=1, value=safe_warning).font = value_font
                sheet.cell(row=current_row, column=1).alignment = left_alignment
                current_row += 1
        else:
            sheet.cell(row=current_row, column=1, value="No warnings").font = value_font
            sheet.cell(row=current_row, column=1).alignment = left_alignment

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 80

    def _row_mapping(self, row: BomRow | Mapping[str, Any]) -> dict[str, Any]:
        if isinstance(row, Mapping):
            return dict(row)
        dumped = row.model_dump()
        return dict(dumped)

    def _exportable_fields(self, rows: list[BomRow]) -> list[str]:
        sample = rows[0].model_dump() if rows else BomRow(project_id=0).model_dump()
        return [field for field in sample.keys() if field != "project"]

    def _sanitize_value(self, value: Any, options: ExportOptions) -> Any:
        value = self._worksheet_safe_value(value)
        if not options.sanitize_formulas:
            return value
        if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
            return "'" + value
        return value

    def _row_warnings(
        self,
        row: Mapping[str, Any],
        columns: Iterable[tuple[str, str]],
    ) -> list[str]:
        warnings: list[str] = []
        for header, field in columns:
            value = row.get(field, "")
            if isinstance(value, str) and self._contains_illegal_characters(value):
                warnings.append(f"{field}: illegal worksheet characters removed")
            if header == "LCSC LINK" and value and not isinstance(value, str):
                warnings.append(f"{field}: non-string hyperlink value converted to text")
            if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
                warnings.append(f"{field}: formula-like text sanitized")
        return warnings

    def _worksheet_safe_value(self, value: Any) -> Any:
        if isinstance(value, datetime) and value.tzinfo is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    def _contains_illegal_characters(self, value: str) -> bool:
        return bool(ILLEGAL_CHARACTERS_RE.search(value))

    def _source_files(self, rows: Sequence[BomRow]) -> str:
        files = sorted({str(row.source_file).strip() for row in rows if str(row.source_file).strip()})
        return ", ".join(files) if files else ""

    def _set_column_widths(
        self,
        sheet,
        rows: Sequence[Mapping[str, Any]],
        headers: Sequence[str],
        sheet_title: str,
    ) -> None:
        for index, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row in rows:
                value = row.get(header, "")
                value_text = "" if value is None else str(value)
                if len(value_text) > max_length:
                    max_length = len(value_text)
            if sheet_title == "BOM":
                width = {
                    "Designator": 20,
                    "Comment": 25,
                    "Footprint": 25,
                    "LCSC LINK": 40,
                    "LCSC PART #": 18,
                }.get(header, min(max(max_length + 2, 12), 40))
            else:
                width = min(max(max_length + 2, 12), 40)
            sheet.column_dimensions[get_column_letter(index)].width = width
